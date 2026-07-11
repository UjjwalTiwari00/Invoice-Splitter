#!/usr/bin/env python3
"""
Invoice Splitter — core logic + Tkinter desktop GUI
Streamlit web app lives in app.py and imports from this module.

CLI usage : python invoice_splitter.py input.xlsx output.xlsx
GUI usage : python invoice_splitter.py
"""

import re
import io
import json
import sys
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HSN_FILE = Path(__file__).parent / "hsn_codes.json"


# ── Constants ────────────────────────────────────────────────────────────────

EMPTY = '—'

# Matches: Product Name x 1.0000 @ 804.24 (Tax 18.0000%: 144.76)
PRODUCT_RE = re.compile(
    r'(.+?)\s+x\s+([\d.]+)\s+@\s+([\d.]+)\s+\(Tax\s+([\d.]+)%:\s+([\d.]+)\)'
)

NUMERIC_COLS = {
    'Quantity', 'Rate', 'Value', 'Gross Total', 'Local Sales',
    'CGST 2.50% Output', 'SGST 2.50% Output',
    'CGST 9% Output', 'SGST 9% Output',
    'IGST 18%', 'Inter State Sales', 'Round Off',
}


# ── Parsing ──────────────────────────────────────────────────────────────────

def parse_particulars(text: str) -> list:
    """
    Extract products from a Particulars cell.
    Format: <Product Name> x <Qty> @ <Rate> (Tax <GST%>: <GST Amount>)
    """
    products = []
    for m in PRODUCT_RE.finditer(text):
        name    = m.group(1).strip().lstrip(',').strip()
        qty     = float(m.group(2))
        rate    = float(m.group(3))
        gst_pct = float(m.group(4))
        gst_amt = float(m.group(5))
        value   = round(qty * rate, 2)
        gross   = round(value + gst_amt, 2)
        products.append({
            'name': name, 'qty': qty, 'rate': rate,
            'value': value, 'gst_pct': gst_pct,
            'gst_amt': gst_amt, 'gross': gross,
        })
    return products


def build_gst_updates(gst_pct: float, gst_amt: float) -> dict:
    """
    GST 5%  → CGST 2.50% + SGST 2.50% (local sales)
    GST 18% → CGST 9%    + SGST 9%    (local sales)
    Other   → IGST column              (inter-state)
    """
    half = round(gst_amt / 2, 3)
    updates = {
        'CGST 2.50% Output': EMPTY,
        'SGST 2.50% Output': EMPTY,
        'CGST 9% Output':    EMPTY,
        'SGST 9% Output':    EMPTY,
        'IGST 18%':          EMPTY,
        'Inter State Sales': EMPTY,
    }
    if gst_pct == 5.0:
        updates['CGST 2.50% Output'] = half
        updates['SGST 2.50% Output'] = half
    elif gst_pct == 18.0:
        updates['CGST 9% Output'] = half
        updates['SGST 9% Output'] = half
    else:
        updates['IGST 18%']          = gst_amt
        updates['Inter State Sales'] = round(gst_amt, 2)
    return updates


# ── HSN code helpers ─────────────────────────────────────────────────────────

def load_hsn_codes() -> dict:
    """Load HSN code map from hsn_codes.json. Keys are upper-cased and stripped."""
    if not HSN_FILE.exists():
        return {}
    raw = json.loads(HSN_FILE.read_text(encoding="utf-8"))
    return {_norm(k): str(v) for k, v in raw.items()}


def save_hsn_codes(hsn_map: dict) -> None:
    """Persist the HSN map (normalised keys) back to hsn_codes.json."""
    HSN_FILE.write_text(
        json.dumps({k: v for k, v in sorted(hsn_map.items())}, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def _norm(text: str) -> str:
    """Normalise a product name for matching: uppercase, collapse whitespace."""
    return ' '.join(str(text).upper().strip().split())


def lookup_hsn(product_name: str, hsn_map: dict) -> str:
    """
    Return the HSN code for *product_name*.
    Strategy (in order):
      1. Exact match (after normalisation)
      2. Lookup key is a substring of the product name
      3. Product name is a substring of a lookup key
    Returns '' if nothing matches.
    """
    norm = _norm(product_name)
    if norm in hsn_map:
        return hsn_map[norm]
    for key, code in hsn_map.items():
        if key in norm or norm in key:
            return code
    return ''


def apply_hsn_codes(df: pd.DataFrame, hsn_map: dict) -> pd.DataFrame:
    """
    Add an 'HSN Code' column to *df* by matching the 'Particulars' column
    against *hsn_map*.  The column is inserted immediately after 'Particulars'.
    """
    df = df.copy()
    norm_map = {_norm(k): str(v) for k, v in hsn_map.items()}
    df['HSN Code'] = df['Particulars'].apply(lambda p: lookup_hsn(str(p), norm_map))

    cols = list(df.columns)
    cols.remove('HSN Code')
    insert_at = cols.index('Particulars') + 1 if 'Particulars' in cols else len(cols)
    cols.insert(insert_at, 'HSN Code')
    return df[cols]


# ── Core processing (shared by GUI and Streamlit) ────────────────────────────

def process_dataframe(df: pd.DataFrame) -> tuple:
    """
    Split every multi-product row in *df* into one row per product.
    Returns (output_df, stats_dict).
    """
    df = df.copy().fillna(EMPTY).astype(str)

    output_rows  = []
    is_split_flags = []
    stats = {
        'input_rows':       len(df),
        'output_rows':      0,
        'split_rows':       0,
        'passthrough_rows': 0,
        'errors':           [],
    }

    for idx, row in df.iterrows():
        particulars = str(row.get('Particulars', EMPTY)).strip()

        if not particulars or particulars == EMPTY:
            output_rows.append(row.to_dict())
            is_split_flags.append(False)
            stats['passthrough_rows'] += 1
            continue

        products = parse_particulars(particulars)

        if not products:
            output_rows.append(row.to_dict())
            is_split_flags.append(False)
            stats['passthrough_rows'] += 1
            stats['errors'].append(
                f"Row {idx + 2}: Could not parse Particulars — kept as-is.\n"
                f"  Value: {particulars[:100]}"
            )
            continue

        stats['split_rows'] += 1
        for p in products:
            r = row.to_dict()
            r.update({
                'Particulars': p['name'],
                'Quantity':    p['qty'],
                'Rate':        p['rate'],
                'Value':       p['value'],
                'Gross Total': p['gross'],
                'Local Sales': p['value'],
                'Round Off':   EMPTY,
            })
            r.update(build_gst_updates(p['gst_pct'], p['gst_amt']))
            output_rows.append(r)
            is_split_flags.append(True)

    out_df = pd.DataFrame(output_rows, columns=df.columns)
    out_df['_is_split'] = is_split_flags   # used for row highlighting in UI

    for col in NUMERIC_COLS:
        if col in out_df.columns:
            try:
                out_df[col] = pd.to_numeric(out_df[col].replace(EMPTY, None))
            except (ValueError, TypeError):
                pass

    stats['output_rows'] = len(out_df)
    return out_df, stats


def dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """Serialise *df* to a formatted Excel file and return the raw bytes."""
    # Drop internal tracking column before export
    export_df = df.drop(columns=['_is_split'], errors='ignore')

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='Split Invoices')
        ws = writer.sheets['Split Invoices']

        # Must set active sheet explicitly — avoids openpyxl "At least one
        # sheet must be visible" IndexError that occurs when Tally exports
        # write phantom formatting across all 16384 Excel columns.
        writer.book.active = ws

        header_fill = PatternFill("solid", fgColor="D9E1F2")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        ws.freeze_panes = 'A2'

        # Only resize the columns that actually have data (not phantom ones)
        data_col_count = len(export_df.columns)
        for i, col_cells in enumerate(ws.columns):
            if i >= data_col_count:
                break
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0)
                for c in col_cells
            )
            ws.column_dimensions[get_column_letter(i + 1)].width = min(max_len + 4, 40)

    buf.seek(0)
    return buf.read()


def process_excel(input_path: str, output_path: str) -> dict:
    """Read *input_path*, process, write formatted Excel to *output_path*."""
    df = pd.read_excel(input_path, dtype=str)
    out_df, stats = process_dataframe(df)
    out_bytes = dataframe_to_excel_bytes(out_df)
    Path(output_path).write_bytes(out_bytes)
    return stats


# ── Desktop GUI ───────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Invoice Splitter — Tally Sales Export")
        self.resizable(False, False)
        self.configure(bg="#f0f4f8")
        self._build_ui()

    def _build_ui(self):
        p = {'padx': 14, 'pady': 7}

        tk.Label(self, text="Invoice Splitter",
                 font=("Segoe UI", 15, "bold"), bg="#f0f4f8", fg="#1a202c"
                 ).grid(row=0, column=0, columnspan=3, pady=(20, 2))
        tk.Label(self,
                 text="Splits multi-product Tally invoice rows into individual product rows",
                 font=("Segoe UI", 9), fg="#718096", bg="#f0f4f8"
                 ).grid(row=1, column=0, columnspan=3, pady=(0, 16))

        tk.Label(self, text="Input File:", bg="#f0f4f8", anchor='w'
                 ).grid(row=2, column=0, sticky='w', **p)
        self.input_var = tk.StringVar()
        tk.Entry(self, textvariable=self.input_var, width=50, state='readonly'
                 ).grid(row=2, column=1, **p)
        tk.Button(self, text="Browse…", command=self._pick_input
                  ).grid(row=2, column=2, **p)

        tk.Label(self, text="Output File:", bg="#f0f4f8", anchor='w'
                 ).grid(row=3, column=0, sticky='w', **p)
        self.output_var = tk.StringVar()
        tk.Entry(self, textvariable=self.output_var, width=50, state='readonly'
                 ).grid(row=3, column=1, **p)
        tk.Button(self, text="Browse…", command=self._pick_output
                  ).grid(row=3, column=2, **p)

        tk.Button(self, text="  Process File  ", command=self._run,
                  bg="#3182ce", fg="white", font=("Segoe UI", 11, "bold"),
                  padx=28, pady=9, relief='flat', cursor='hand2'
                  ).grid(row=4, column=0, columnspan=3, pady=16)

        tk.Label(self, text="Log:", bg="#f0f4f8", anchor='w'
                 ).grid(row=5, column=0, sticky='w', padx=14)
        self.log_box = tk.Text(self, height=12, width=72, state='disabled',
                               font=("Consolas", 9), bg="#1a202c", fg="#e2e8f0",
                               insertbackground='white', relief='flat')
        self.log_box.grid(row=6, column=0, columnspan=3, padx=14, pady=(2, 18))

    def _pick_input(self):
        path = filedialog.askopenfilename(
            title="Select Input Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if path:
            self.input_var.set(path)
            p = Path(path)
            self.output_var.set(str(p.parent / (p.stem + "_split" + p.suffix)))

    def _pick_output(self):
        path = filedialog.asksaveasfilename(
            title="Save Output As", defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.output_var.set(path)

    def _log(self, msg: str):
        self.log_box.configure(state='normal')
        self.log_box.insert('end', msg + '\n')
        self.log_box.see('end')
        self.log_box.configure(state='disabled')
        self.update_idletasks()

    def _run(self):
        inp = self.input_var.get().strip()
        out = self.output_var.get().strip()
        if not inp:
            messagebox.showwarning("Missing Input", "Please select an input Excel file.")
            return
        if not out:
            messagebox.showwarning("Missing Output", "Please choose an output file path.")
            return

        self._log(f"Input  : {inp}")
        self._log(f"Output : {out}")
        self._log("Processing...")
        try:
            s = process_excel(inp, out)
            self._log(f"Input rows       : {s['input_rows']}")
            self._log(f"Output rows      : {s['output_rows']}")
            self._log(f"Rows split       : {s['split_rows']}")
            self._log(f"Rows passed thru : {s['passthrough_rows']}")
            if s['errors']:
                self._log(f"\nWarnings ({len(s['errors'])}):")
                for e in s['errors']:
                    self._log(f"  ! {e}")
            self._log(f"\nDone! Saved to: {out}")
            self._log("─" * 60)
            messagebox.showinfo("Complete",
                                f"Processing complete!\n\n"
                                f"Input rows:   {s['input_rows']}\n"
                                f"Output rows:  {s['output_rows']}\n\n"
                                f"Saved to:\n{out}")
        except Exception as exc:
            self._log(f"ERROR: {exc}")
            messagebox.showerror("Error", str(exc))
            raise


# ── CLI / entry point ─────────────────────────────────────────────────────────

def main():
    if len(sys.argv) == 3:
        inp, out = sys.argv[1], sys.argv[2]
        print(f"Processing: {inp}")
        s = process_excel(inp, out)
        print(f"Input rows       : {s['input_rows']}")
        print(f"Output rows      : {s['output_rows']}")
        print(f"Rows split       : {s['split_rows']}")
        print(f"Rows passed thru : {s['passthrough_rows']}")
        if s['errors']:
            print("\nWarnings:")
            for e in s['errors']:
                print(f"  {e}")
        print(f"\nSaved to: {out}")
    else:
        App().mainloop()


if __name__ == '__main__':
    main()
